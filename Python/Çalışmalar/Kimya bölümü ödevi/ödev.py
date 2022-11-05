from openpyxl import Workbook,load_workbook
names,CS102,CHEM102,CHEM112,CHEM132,ENG102,PHYS102,PHYS112,MATH142=[],[],[],[],[],[],[],[],[]
students = {}
def  extract_elements(a):
    wb = load_workbook(a)
    ws = wb.active
    for satir in range(2,ws.max_row+1):
        for sutun in range(2,ws.max_column):
            if sutun==2:
                names.append(str(ws.cell(satir,sutun).value))
            elif sutun==3:
                CS102.append(str(ws.cell(satir,sutun).value))
            elif sutun==4:
                CHEM102.append(str(ws.cell(satir,sutun).value))
            elif sutun==5:
                CHEM112.append(str(ws.cell(satir,sutun).value))
            elif sutun==6:
                CHEM132.append(str(ws.cell(satir,sutun).value))
            elif sutun==7:
                ENG102.append(str(ws.cell(satir,sutun).value))
            elif sutun==8:
                PHYS102.append(str(ws.cell(satir,sutun).value))
            elif sutun==9:
                PHYS112.append(str(ws.cell(satir,sutun).value))
            elif sutun==10:
                MATH142.append(str(ws.cell(satir,sutun).value))
    wb.close()
    return [CS102,CHEM102,CHEM112,CHEM132,ENG102,PHYS102,PHYS112,MATH142],names         


def convert_grades_to_letters(notes):
    for e in range(0,15):
        for q in notes:
    
            if int(q[e]) >=90:
                q[e]="AA" 
            elif int(q[e]) >=80 and int(q[e])<90:
                q[e]="BA"   
            elif int(q[e]) >=70 and int(q[e])<80:
                q[e]="BB" 
            elif int(q[e]) >=60 and int(q[e])<70:
                q[e]="BC"
            elif int(q[e]) >=50 and int(q[e])<60:
                q[e]="CC" 
            elif int(q[e]) >=40 and int(q[e])<50:
                q[e]="DC"        
            elif int(q[e]) >=30 and int(q[e])<40:
                q[e]="DD"         
            elif int(q[e]) >= 20 and int(q[e])<30:
                q[e]="DF"
            else:
                q[e]="FF"
                


def calculate_gpa():
    akts = [3,5,3,2,3,6,2,6]
    for e in range(0,15): 
        gpa = 0
        for q in notes:
            credit = akts[notes.index(q)] 
    
            if q[e]=="AA":
                point=(4*credit)
                gpa += point
            elif q[e]=="BA":
                gpa += (3.5*credit)
            elif q[e]=="BB":
                gpa += (3*credit)
            elif q[e]=="BC":
                gpa += (2.5*credit)
            elif q[e]=="CC":
                gpa += (2*credit)
            elif q[e]=="DC":
                gpa += (1.5*credit)
            elif q[e]=="DD":
                gpa += (1*credit)
            elif q[e]=="DF":
                gpa += (0.5*credit)
            else:
                gpa+=0*credit
            
                    
        gpa = gpa/30
        students[names[e]] = gpa
        
def prepare_letters_and_av_gpa():
    wy = Workbook()
    wq = wy.active
    wq['A1'] = "STUDENTS"
    wq['B1'] = "CS102"
    wq['C1'] = "CHEM102"
    wq['D1'] = "CHEM112"
    wq['E1'] = "CHEM132"
    wq['F1'] = "ENG102"
    wq['G1'] = "PHYS102"
    wq['H1'] = "PHYS112"
    wq['I1'] = "MATH142"
    wq['J1'] = "GPA"
    for a in range(150):
        case = ""
        if a<14:
            case = "A"
            cell = case+str((a%15)+2)
            wq[cell] = names[(a%15)]
        elif a<29 and a>14:
            case = "B"
            cell = case+str((a%15)+2)
            wq[cell] = notes[0][(a%15)]
        elif a<44 and a>29:
            case = "C"
            cell = case+str((a%15)+2)
            wq[cell] = notes[1][(a%15)]            
        elif a<59 and a>44:
            case = "D"
            cell = case+str((a%15)+2)
            wq[cell] = notes[2][(a%15)]   
        elif a<74 and a>59:
            case = "E"
            cell = case+str((a%15)+2)
            wq[cell] = notes[3][(a%15)]   
        elif a<89 and a>74:
            case = "F"
            cell = case+str((a%15)+2)
            wq[cell] = notes[4][(a%15)]   
        elif a<104 and a>89:
            case = "G"
            cell = case+str((a%15)+2)
            wq[cell] = notes[5][(a%15)]   
        elif a<119 and a>104:
            case = "H"
            cell = case+str((a%15)+2)
            wq[cell] = notes[6][(a%15)]   
        elif a<134 and a>119:
            case = "I"
            cell = case+str((a%15)+2)
            wq[cell] = notes[7][(a%15)]   
        elif a<149 and a>134:
            case = "J"
            cell = case+str((a%15)+2)
            wq[cell] = students[names[(a%15)]] 
              
    wy.save("end_term_grades.xlsx")
    
notes,names=extract_elements("students_grades.xlsx")    
convert_grades_to_letters(notes)                   
calculate_gpa()           
prepare_letters_and_av_gpa()
print(students)
    
    
            
            
            
            
